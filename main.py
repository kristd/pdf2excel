# This is a sample Python script.
#import openpyxl
# Press ⌃R to execute it or replace it with your code.
# Press Double ⇧ to search everywhere for classes, files, tool windows, actions, and settings.


import pdfplumber
import pandas as pd
import openpyxl
import os
import glob
import datetime
import re
import calendar
import shutil
from loguru import logger
import json



def getConfig(filename):
    try:
        f = open(filename, 'r',encoding='utf-8')
    except FileNotFoundError:
        logger.debug(filename + ' is not found!!!!!')
    content = f.read()
    a = json.loads(content)
    return a

def rename_file_from_season(folder_path, file_name_pattern):
    try:
        for old_f in sorted(os.listdir(folder_path), reverse=True):
            if re.search(file_name_pattern, old_f) is not None:
                if re.search(r'_\d+.xlsx', old_f) is None:
                    logger.debug('Moving file from ' + os.path.join(folder_path,old_f) + ' to ' + os.path.join(folder_path,str(
                        orderNum) + '_Single_Sheet_1.xlsx'))
                    shutil.move(os.path.join(folder_path,old_f), os.path.join(folder_path,str(
                        orderNum) + '_Single_Sheet_1.xlsx'))
                else:
                    index = int(old_f.split('_')[3].split('.')[0]) + 1
                    logger.debug('Moving file from ' + os.path.join(folder_path,old_f) + ' to ' + os.path.join(folder_path,str(
                        orderNum) + '_Single_Sheet_' + str(index) + '.xlsx'))
                    shutil.move(os.path.join(folder_path,old_f), os.path.join(folder_path,str(
                        orderNum) + '_Single_Sheet_' + str(index) + '.xlsx'))
    except InterruptedError:
        return 0
    return 1


def get_purchase_order_first_page_text(i):
    try:
        pdf_object = pdfplumber.open(i)
    except InterruptedError:
        logger.debug(i + ' is unable to be opened....')
    return pdf_object


def get_sizecolourbreakdown_pages_text(i):
    try:
        pdf_object = pdfplumber.open(i)
    except InterruptedError:
        logger.debug(i + ' is unable to be opened....')
    return pdf_object


def get_file_validation(file_path, orderNum, flag):
    if flag == 1:
        f = glob.glob(file_path + str(orderNum) + '_SizePerColourBreakdown*')
        if len(f) > 0:
            return True
        else:
            return False
    else:
        f = glob.glob(file_path + 'updated_' + str(orderNum) + '_SizePerColourBreakdown*')
        if len(f) > 0:
            return True
        else:
            return False


def split_lines(text):
    return text.splitlines()


def get_brand(text):
    return text


def get_order_type(text):
    if text.find('Online') == -1:
        return 'Store'
    else:
        return 'Online'


def get_season(text):
    return 'S' + re.findall(r'\d+-\d+', text)[0]


def get_order_no(text):
    return re.findall(r'\d+', text)[0]


def get_department_no(text):
    return re.findall(r'\d+', text)[1]


def get_product_no(text):
    return re.findall(r'\d+', text)[2]


def get_date_of_order(text):
    dateOfOrder = re.findall(r'\d+ [a-zA-Z]+, \d{4}', text)[0]
    l_day = re.findall(r'\d{2} ', dateOfOrder)[0].strip()
    l_mon = list(calendar.month_abbr).index(re.findall(r'[a-zA-Z]{3},', dateOfOrder)[0].replace(',', ''))
    l_year = re.findall(r'\d{4}', dateOfOrder)[0]
    return str(l_year) + '-' + str(l_mon) + '-' + str(l_day)


def get_product_name(text):
    return re.findall(r'Product Name: .*', text)[0].split(':')[1].strip()


def get_development_no(text):
    return re.findall(r'Development .*', text)[0].split('  ')[1]


def get_no_of_pieces(text):
    return re.findall(r'\d+', text)[0]


def get_detail_country_code(text):
    country_code = re.findall(r' [a-zA-Z]{2} \(', text)[0].replace(')', '').replace('(', '').strip()
    return country_code


def get_delivery_dates_dicts(text, p1, p2):
    time_delivery_dict = {}
    time_delivery_dict.clear()
    for p in range(p1, p2):
        logger.debug(text[p])
        if re.search(r', \d{4}', text[p]) is not None:
            l_day = text[p][0:2]
            l_mon = list(calendar.month_abbr).index(text[p][3:6])
            l_year = text[p][8:12]
            logger.debug('p1: ' + str(p1) + ', p2: ' + str(p2))
            logger.debug('p: ' + str(p) + ', p+1: ' + str(p+1) + ', p+2: ' + str(p+2))
            logger.debug('p: ' + text[p] + ', p+1: ' + text[p+1] +', p+2: ' + text[p+2])
            if p == p2-1:  ###last line
                time_delivery_dict[re.sub(u"\\(.*?\\)", '', text[p].replace('\xa0', '')[13:].replace(
                    re.findall(r'\d+ .*%', text[p].replace('\xa0', '')[13:])[0], '').strip())] = str(
                    l_year) + '-' + str(l_mon) + '-' + str(l_day)
            if (p + 1) <= p2 and re.search(r', \d{4}', text[p + 1]) is not None:  ### 1 TOD only 1 line case
                time_delivery_dict[re.sub(u"\\(.*?\\)", '', text[p].replace('\xa0', '')[13:].replace(
                    re.findall(r'\d+ .*%', text[p].replace('\xa0', '')[13:])[0], '').strip())] = str(
                    l_year) + '-' + str(l_mon) + '-' + str(l_day)
            elif ((p+2) <= p2 and re.search(r', \d{4}', text[p + 2]) is not None) or re.search(r'Total',text[p+2]) is not None:  #### 1 TOD 2 lines case
                time_delivery_dict[re.sub(u"\\(.*?\\)", '', (text[p].replace('\xa0', '')[13:].replace(
                    re.findall(r'\d+ .*%', text[p].replace('\xa0', '')[13:])[0], '').strip() +
                                                             text[p + 1].replace('\xa0', '').strip()))] = str(
                    l_year) + '-' + str(
                    l_mon) + '-' + str(l_day)
            elif (p+2) <= p2 and re.search(r', \d{4}', text[p + 2]) is None:  #### 1 TOD 3 lines case
                time_delivery_dict[re.sub(u"\\(.*?\\)", '', (text[p].replace('\xa0', '')[13:].replace(
                    re.findall(r'\d+ .*%', text[p].replace('\xa0', '')[13:])[0], '').strip() +
                                                             text[p + 1].replace('\xa0', '').strip()+ text[p+2].replace('\xa0', '').strip()))] = str(
                    l_year) + '-' + str(
                    l_mon) + '-' + str(l_day)
    return time_delivery_dict


def get_price_dicts(text, p1, p2):
    price_dict = {}
    price_dict.clear()
    for p in range(p1, p2):
        if re.search(r'\d+', text[p]) is not None:
            if (p + 1) <= p2 and re.search(r'\d+', text[p + 1]) is not None:
                l_price = text[p].split(' ')[0]
                l_currency = text[p].split(' ')[1]
                l_countries = text[p].replace(text[p].split(' ')[0], '').replace(
                    text[p].split(' ')[1], '').strip()
                price_dict[l_countries] = str(l_price) + '-' + str(l_currency)
            else:
                l_price = text[p].split(' ')[0]
                l_currency = text[p].split(' ')[1]
                l_countries = text[p].replace(text[p].split(' ')[0], '').replace(
                    text[p].split(' ')[1], '').strip() + text[p + 1].strip()
                price_dict[l_countries] = str(l_price) + '-' + str(l_currency)
    return price_dict


def get_term_dicts(text, p1, p2, country_code):
    logger.debug('in get_term_dicts')
    term_dicts = {}
    term_dicts.clear()

    for p in range(p1, p2):
        country_list = text[p].split(',')
        for cl in country_list:
            if re.search(r'/', cl.strip()) is not None:
                for c in country_codes:
                    if re.search(cl.strip(), c) is not None:
                        term_dicts[c[3:]] = text[p + 1].split(' ')[2]
                        break
            else:
                for c in country_codes:
                    if re.search(cl.strip(), c) is not None:
                        term_dicts[c] = text[p + 1].split(' ')[2]
                        break
    return term_dicts


def get_colourname_dicts(text, p1, p2):
    colourname_dicts = {}
    colourname_dicts.clear()
    for p in range(p1, p2):
        colourname_dicts[text[p].split(' ')[0]] = re.findall(r'[a-zA-Z]+.*[a-zA-Z]+',
                                                             text[p].replace('Solid', '').replace('USD', '').replace(
                                                                 'All over pattern', ''))[0]
    return colourname_dicts


def write_data_into_summary_excel(file_path, data, season):
    file_name = 'Orders_Summary_' + season + '.xlsx'
    if not os.path.exists(file_path + file_name):
        workbook = openpyxl.Workbook()
        sheet = workbook.create_sheet('Summary', index=0)
        header = ["品牌/Brands", "订单类型/Order Type", "季度/Season", "订单号/Order No", "部门号/Department No",
                  "生产货号/Product No", "下单日期/Date of Order", "产品名称/Product Name",
                  "开发货号/Development No",
                  "双/包/No of Pieces", "国家/Country", "运输方式/Terms of Delivery",
                  "HM色号+颜色描述/Colour Code & Colour Name", "尺码/Size", "装箱方式/Packing Type","配比/Assortment","数量/No of Assortment",
                  "订单数量/Qty/Artical", "单价/Cost", "货币/Currency",
                  "总数量/总件数/Total Pairs/pcs", "出货日期/TOD", "编号/Artical No", "下载日期/Download Date"]
        sheet.append(header)
        for single_row in data:
            sheet.append(single_row)
        workbook.save(file_path + file_name)
        workbook.close()
    else:
        wb = openpyxl.load_workbook(file_path + file_name)
        wb_sheet_name = wb.get_sheet_by_name(wb.get_sheet_names()[0])
        for single_row in data:
            wb_sheet_name.append(single_row)
        wb.save(file_path + file_name)
        wb.close()


def write_data_into_excel(file_path, season, orderNum, data):
    #######write data into excel sheet
    ##read excel file , if file not existed, create a new one.
    if not os.path.exists(file_path + 'power_bi'):
        os.makedirs(file_path + 'power_bi')
    if os.path.exists(file_path + season):
        os.chdir(file_path + season)
        if not os.path.exists(os.path.join(file_path,season,str(
                orderNum) + '_Single_Sheet.xlsx')):
            workbook = openpyxl.Workbook()
            sheet = workbook.create_sheet(orderNum, index=0)
            header = ["品牌/Brands", "订单类型/Order Type", "季度/Season", "订单号/Order No", "部门号/Department No",
                      "生产货号/Product No", "下单日期/Date of Order", "产品名称/Product Name",
                      "开发货号/Development No",
                      "双/包/No of Pieces", "国家/Country", "运输方式/Terms of Delivery",
                      "HM色号+颜色描述/Colour Code & Colour Name", "尺码/Size", "装箱方式/Packing Type","配比/Assortment","数量/No of Assortment",
                      "订单数量/Qty/Artical", "单价/Cost", "货币/Currency",
                      "总数量/总件数/Total Pairs/pcs", "出货日期/TOD", "编号/Artical No", "下载日期/Download Date"]
            sheet.append(header)
            for single_row in data:
                sheet.append(single_row)
            workbook.save(os.path.join(file_path,season,str(
                orderNum) + '_Single_Sheet.xlsx'))
            logger.debug('power_bi file: ' + os.path.join(file_path, 'power_bi', str(orderNum) + '_Single_Sheet.xlsx'))
            workbook.save(os.path.join(file_path, 'power_bi', str(orderNum) + '_Single_Sheet.xlsx'))
            workbook.close()
            ##write_data_into_summary_excel(file_path, data, season)
        else:
            ##append data
            wb = openpyxl.load_workbook(os.path.join(file_path,season,str(orderNum) + '_Single_Sheet.xlsx'))
            wb_sheet_name = wb.get_sheet_by_name(wb.get_sheet_names()[0])
            for single_row in data:
                wb_sheet_name.append(single_row)
            wb.save(os.path.join(file_path,season,str(orderNum) + '_Single_Sheet.xlsx'))
            logger.debug('power_bi file: ' + os.path.join(file_path, 'power_bi', str(orderNum) + '_Single_Sheet.xlsx'))
            wb.save(os.path.join(file_path, 'power_bi', str(orderNum) + '_Single_Sheet.xlsx'))
            wb.close()
            ###Summary Sheet
            ##write_data_into_summary_excel(file_path, data, season)
    else:
        os.makedirs(file_path + season)
        os.chdir(file_path + season)
        if not os.path.exists(os.path.join(file_path,season,str(orderNum) + '_Single_Sheet.xlsx')):
            workbook = openpyxl.Workbook()
            sheet = workbook.create_sheet(orderNum, index=0)
            header = ["品牌/Brands", "订单类型/Order Type", "季度/Season", "订单号/Order No", "部门号/Department No",
                      "生产货号/Product No", "下单日期/Date of Order", "产品名称/Product Name",
                      "开发货号/Development No",
                      "双/包/No of Pieces", "国家/Country", "运输方式/Terms of Delivery",
                      "HM色号+颜色描述/Colour Code & Colour Name", "尺码/Size", "装箱方式/Packing Type","配比/Assortment","数量/No of Assortment",
                      "订单数量/Qty/Artical", "单价/Cost", "货币/Currency",
                      "总数量/总件数/Total Pairs/pcs", "出货日期/TOD", "编号/Artical No", "下载日期/Download Date"]
            sheet.append(header)
            for single_row in data:
                sheet.append(single_row)
            workbook.save(os.path.join(file_path,season,str(
                orderNum) + '_Single_Sheet.xlsx'))
            logger.debug('power_bi file: ' + os.path.join(file_path, 'power_bi', str(orderNum) + '_Single_Sheet.xlsx'))
            workbook.save(os.path.join(file_path, 'power_bi', str(orderNum) + '_Single_Sheet.xlsx'))  ###power bi
            workbook.close()
            ##write_data_into_summary_excel(file_path, data, season)

        else:
            ##append data
            wb = openpyxl.load_workbook(os.path.join(file_path,season,str(
                orderNum) + '_Single_Sheet.xlsx'))
            wb_sheet_name = wb.get_sheet_by_name(wb.get_sheet_names()[0])
            for single_row in data:
                wb_sheet_name.append(single_row)
            wb.save(os.path.join(file_path,season,str(
                orderNum) + '_Single_Sheet.xlsx'))
            logger.debug('power_bi file: ' + os.path.join(file_path, 'power_bi', str(orderNum) + '_Single_Sheet.xlsx'))
            wb.save(os.path.join(file_path, 'power_bi', str(
                orderNum) + '_Single_Sheet.xlsx'))
            wb.close()
            ###Summary Sheet
            ##write_data_into_summary_excel(file_path, data, season)


# with pdfplumber.open('/Users/kristd/Downloads/2PDF/667098_PurchaseOrder_20230627_144321.pdf') as pdf:
#     for page in pdf.pages:
#         print(page.extract_text_simple())
#         # 每页打印一分页分隔
#         print('---------- 分页分隔 ----------')


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    path_dicts = getConfig('config.json')
    log_path = path_dicts['main_path']
    logger.debug('log_path: ' + log_path)
    ###initial log file
    log_file = logger.add(log_path + 'pdf2excel.log')
    # get PurchaseOrder files list
    formatted_today = datetime.date.today().strftime('%Y%m%d')
    file_path = path_dicts['pdf_order_path']
    f = glob.glob(file_path + '*PurchaseOrder*')
    pattern = 'UPDATED_*'
    excel_file_path = path_dicts['excel_order_path']
    code_file_path = path_dicts['main_path']
    df = pd.read_excel(code_file_path + 'country_code.xlsx')
    df_li = df.values.tolist()
    country_codes = []
    for s_li in df_li:
        country_codes.append(s_li[0] + '-' + s_li[1])
    for i in f:  # PurchaseOrder loop, the outer loop
        data = []
        file_name = os.path.basename(i)
        if re.search(pattern, file_name.upper()) is None:
            orderNum = file_name.split('_', 1)[0]
            logger.debug('Current Order Number: ' + str(orderNum))
            try:
                if get_file_validation(file_path, orderNum, 1):
                    # get the PurchaseOrder info and go down to the detail loop!!!!!!!!!!
                    logger.debug('----------------------- New Order Scenario -----------------------')
                    ####New Order Scenario
                    order_object = get_purchase_order_first_page_text(i)
                    order_text = order_object.pages[0].extract_text_simple()
                    order_info_array = split_lines(order_text)
                    brand = get_brand(order_info_array[1])  # Column A
                    order_type = get_order_type(order_info_array[0])  # Column B
                    season = get_season(order_info_array[6])  # Column C
                    order_no = get_order_no(order_info_array[3])  # Column D
                    department_no = get_department_no(order_info_array[3])  # Column E
                    product_no = get_product_no(order_info_array[3])  # Column F
                    date_of_order = get_date_of_order(order_info_array[5])  # Column G
                    product_name = get_product_name(order_info_array[4])  # Column H
                    development_no = get_development_no(order_info_array[9])  # Column I
                    no_of_pieces = get_no_of_pieces(order_info_array[12])  # Column J
                    download_date = datetime.date.today().strftime('%Y-%m-%d')  # Column W

                    ### create the price/country mapping information
                    terms_1st_position = 0
                    terms_last_position = 0
                    time_delivery_1st_position = 0
                    time_delivery_last_position = 0
                    price_1st_position = 0
                    price_last_position = 0
                    colourname_1st_position = 0
                    colourname_last_position = 0
                    for p in range(0, len(order_info_array)):
                        if re.search(r'Terms of Delivery', order_info_array[p]) is not None:
                            terms_1st_position = p
                        if re.search(r'Time of Delivery Planning Markets', order_info_array[p]) is not None:
                            terms_last_position = p - 1
                            time_delivery_1st_position = p
                        if re.search(r'Total: \d+', order_info_array[p]) is not None:
                            time_delivery_last_position = p - 1
                        if re.search(r'Invoice Average Price', order_info_array[p]) is not None:
                            price_1st_position = p
                        if (re.search(r'By accepting and performing under this Order, the Supplier acknowledges:', order_info_array[p]) is not None) or (re.search(r'Please note that this order concerns a licensed product.*', order_info_array[p]) is not None):
                            if price_last_position == 0 and re.search(r'By accepting and performing under this Order, the Supplier acknowledges:', order_info_array[p]) is not None:
                                price_last_position = p - 1
                            elif price_last_position > p and re.search(r'By accepting and performing under this Order, the Supplier acknowledges:', order_info_array[p]) is not None:
                                price_last_position = p - 1
                            elif price_last_position == 0 and re.search(r'Please note that this order concerns a licensed product.*', order_info_array[p]) is not None:
                                price_last_position = p - 2
                            elif price_last_position > p and re.search(r'Please note that this order concerns a licensed product.*', order_info_array[p]) is not None:
                                price_last_position = p - 2
                        if (re.search(r'Article No H&M Colour Code', order_info_array[p])):
                            colourname_1st_position = p
                        if (re.search(r'Total Quantity:', order_info_array[p])):
                            colourname_last_position = p - 1


                    logger.debug('terms_1st_position: ' + str(terms_1st_position))
                    logger.debug('terms_last_position: ' + str(terms_last_position))
                    logger.debug('time_delivery_1st_position: ' + str(time_delivery_1st_position))
                    logger.debug('time_delivery_last_position: ' + str(time_delivery_last_position))
                    logger.debug('price_1st_position: ' + str(price_1st_position))
                    logger.debug('price_last_position: ' + str(price_last_position))
                    logger.debug('colourname_1st_position: ' + str(colourname_1st_position))
                    logger.debug('colourname_last_position: ' + str(colourname_last_position))
                    ### create the country/term mapping
                    logger.debug('getting term dict...')
                    term_dict = {}
                    term_dict.clear()
                    logger.debug('getting country code file....')
                    term_dict = get_term_dicts(order_info_array, terms_1st_position + 1, terms_last_position + 1,country_codes)

                    ##create the time delivery mapping
                    logger.debug('getting time delivery dict...')
                    time_delivery_dict = {}
                    time_delivery_dict.clear()
                    time_delivery_dict = get_delivery_dates_dicts(order_info_array, time_delivery_1st_position + 1,
                                                                  time_delivery_last_position + 1)
                    logger.debug('Print time Delivery Dict: ')
                    logger.debug(time_delivery_dict)

                    ##create price mapping
                    logger.debug('getting price dict...')
                    price_dict = {}
                    price_dict = get_price_dicts(order_info_array, price_1st_position + 1, price_last_position + 1)
                    logger.debug(price_dict)

                    ##create colourname mapping
                    colourname_dict = {}
                    colourname_dict.clear()
                    colourname_dict = get_colourname_dicts(order_info_array, colourname_1st_position + 1,
                                                           colourname_last_position + 1)

                    ###detail loop start here
                    detail_pages = []
                    detail_obj = get_sizecolourbreakdown_pages_text(
                        i.replace('PurchaseOrder', 'SizePerColourBreakdown'))
                    detail_pages = detail_obj.pages
                    for dp in range(0, len(detail_pages)):
                        country_name = ''  # Column K
                        fright_term = ''  # Column L
                        cost = ''  # Column Q
                        currency = ''  # Column R
                        TOD = ''  # Column T

                        page_text = detail_pages[dp].extract_text_simple()
                        order_detail_array = split_lines(page_text)

                        assortment_1st_position = 0
                        assortment_last_position = 0
                        solid_1st_position = 0
                        solid_last_position = 0

                        for p in range(0, len(order_detail_array)):
                            if (re.search(r'Assortment', order_detail_array[p])):
                                assortment_1st_position = p
                            if (re.search(r'Solid', order_detail_array[p])):
                                solid_1st_position = p
                                assortment_last_position = p - 1
                            if (re.search(r'Total', order_detail_array[p])):
                                solid_last_position = p - 1

                        country_code = get_detail_country_code(order_detail_array[11])
                        ## get term

                        for k in term_dict.keys():
                            if re.search(country_code, str(k)) is not None or re.search(str(k), country_code) is not None:
                                country_name = k
                                fright_term = term_dict[k]
                                break
                        ## get cost and currency
                        for k in price_dict.keys():
                            if re.search('/',country_name) is not None:
                                if re.search(country_name.split('/')[0],str(k)) is not None:
                                    cost = price_dict[k].split('-')[0]
                                    currency = price_dict[k].split('-')[1]
                                    break
                            if re.search(country_code, str(k)) is not None:
                                cost = price_dict[k].split('-')[0]
                                currency = price_dict[k].split('-')[1]
                                break

                        for k in time_delivery_dict.keys():
                            if re.search('/',country_name) is not None:
                                if re.search(country_name.split('/')[0],str(k)) is not None:
                                    TOD = time_delivery_dict[k]
                                    break
                            if re.search(country_code, str(k)) is not None or re.search(str(k), country_name) is not None:
                                TOD = time_delivery_dict[k]
                                break

                        ##get artical NO & colourcode+colourname
                        artical_list = re.findall(r'\d+.*\d+', order_detail_array[12])[0].split(' ')
                        colourcode_list = re.findall(r'\d+-.*', order_detail_array[13])[0].split(' ')
                        # print(country_name)
                        for a in artical_list:
                            artical_no = ''  # Column U
                            colourcode_colourname = ''  # Column M
                            size = ''  # Column N
                            packing_type = ''  # Column O
                            l_assortment_qty = '' #Column P
                            l_no_assortment_qty = ''
                            qty_artical = ''  # Column Q
                            artical_no = a
                            for k in colourname_dict.keys():
                                if a == k:
                                    colourcode_colourname = colourcode_list[artical_list.index(a)] + ' ' + str(
                                        colourname_dict[a])
                                    break
                            ## assorment/solid qty information

                            if assortment_1st_position > 0:
                                packing_type = 'Assortment'
                                # print(packing_type)
                                no_of_asst_list = []
                                for ap in range(assortment_1st_position + 1, assortment_last_position + 1):
                                    if re.search(r'No of Asst:', order_detail_array[ap]) is not None:
                                        no_of_asst_list = order_detail_array[ap].replace('No of Asst: ', '').split(' ')
                                        break
                                for ap in range(assortment_1st_position + 1, assortment_last_position + 1):
                                    if re.search(r'\(.*\)\*', order_detail_array[ap]) is not None:
                                        size = re.findall(r'\(.*\)\*', order_detail_array[ap])[0].replace(')*',
                                                                                                          '').replace(
                                            '(', '')
                                        if re.findall(r'\* \d+.*', order_detail_array[ap]) == []:
                                            next
                                        else:
                                            if len(re.findall(r'\* \d+.*', order_detail_array[ap])[0].replace('* ',
                                                                                                              '').split(
                                                    ' ')) >= (artical_list.index(a)+1):
                                                if re.findall(r'\* \d+.*', order_detail_array[ap]) is not [] and int(
                                                    re.findall(r'\* \d+.*', order_detail_array[ap])[0].replace('* ',
                                                                                                               '').split(
                                                        ' ')[artical_list.index(a)]) > 0:
                                                    l_assortment_qty = int(
                                                    re.findall(r'\* \d+.*', order_detail_array[ap])[0].replace('* ',
                                                                                                               '').split(
                                                        ' ')[artical_list.index(a)])
                                                    qty_artical = int(
                                                    re.findall(r'\* \d+.*', order_detail_array[ap])[0].replace('* ',
                                                                                                               '').split(
                                                        ' ')[artical_list.index(a)]) * int(
                                                    no_of_asst_list[artical_list.index(a)])
                                                    l_no_assortment_qty = int(
                                                    no_of_asst_list[artical_list.index(a)])
                                                ####call write_data_into_excel()
                                                ##prepare dataset
                                                    data.append([brand, order_type, season, order_no, department_no, product_no,
                                                        date_of_order, product_name, development_no, no_of_pieces,
                                                        country_name, fright_term.replace(',', '')
                                                    , colourcode_colourname, size, packing_type, l_assortment_qty, l_no_assortment_qty,qty_artical, cost,
                                                        currency, int(no_of_pieces) * int(qty_artical), TOD, artical_no,
                                                        download_date])
                                                    logger.debug('Appending data into dateaset... ')

                            if solid_1st_position > 0:
                                packing_type = 'Solid'
                                # print(packing_type)
                                for sp in range(solid_1st_position + 1, solid_last_position + 1):
                                    if re.search(r'\(.*\)\*', order_detail_array[sp]) is not None:
                                        size = re.findall(r'\(.*\)\*', order_detail_array[sp])[0].replace(')*',
                                                                                                          '').replace(
                                            '(', '')
                                        # print(re.findall(r'\* \d+.*', order_detail_array[sp])[0].replace('* ',''))
                                        if re.findall(r'\* \d+.*', order_detail_array[sp]) == []:
                                            next
                                        else:
                                            ##print('size: ' + str(len(re.findall(r'\* \d+.*', order_detail_array[sp])[0].replace('* ','').split(' ')))+', ' + 'current index: ' + str(artical_list.index(a)))
                                            if len(re.findall(r'\* \d+.*', order_detail_array[sp])[0].replace('* ','').split(' ')) >= (artical_list.index(a) +1):
                                                if re.findall(r'\* \d+.*', order_detail_array[sp]) is not [] and int(re.findall(r'\* \d+.*', order_detail_array[sp])[0].replace('* ','').split(' ')[artical_list.index(a)]) > 0:
                                                    l_assortment_qty = ''
                                                    l_no_assortment_qty = ''
                                                    qty_artical = int(
                                                        re.findall(r'\* \d+.*', order_detail_array[sp])[0].replace('* ',
                                                                                                               '').split(' ')[artical_list.index(a)])
                                                #######write data into excel sheet
                                                ##read excel file , if file not existed, create a new one.
                                                # prepare dataset
                                                    data.append([brand, order_type, season, order_no, department_no, product_no,
                                                        date_of_order, product_name, development_no, no_of_pieces,
                                                        country_name, fright_term.replace(',', ''),
                                                        colourcode_colourname, size, packing_type,l_assortment_qty,l_no_assortment_qty, qty_artical, cost,
                                                        currency, int(no_of_pieces) * int(qty_artical), TOD, artical_no,
                                                        download_date])
                                                ##write_data_into_excel(excel_file_path, season, orderNum, data)
                    detail_obj.close()
                    order_object.close()
                    logger.debug('I am here..after order object closed...')
                    ######end of detail loop
                else:
                    # throw error as the child file is not exists.
                    try:
                        f = open(file_path + str(orderNum) + '_SizePerColourBreakdown*',encoding='utf-8',errors='ignore')
                    except FileNotFoundError:
                        logger.debug(file_path + str(orderNum) + ', SizePerColourBreakdown is not found!!!!')
            except FileNotFoundError:
                logger.error('OrderNum: ' + str(orderNum) + ', file is not found!')

            logger.debug('Writing data into excel sheet....')
            write_data_into_excel(excel_file_path, season, orderNum, data)
            ###move processed file into Archive folder.
            try:
                if os.path.exists(i.replace('PurchaseOrder', 'SizePerColourBreakdown')):
                    ##get the PurchaseOrder info and go down to the detail loop!!!!!!!!!!
                    ####Update Order Scenarios
                    archive_path = path_dicts['archive_path']
                    shutil.move(i, archive_path + file_name)
                    shutil.move(i.replace('PurchaseOrder', 'SizePerColourBreakdown'),
                                archive_path + file_name.replace('PurchaseOrder', 'SizePerColourBreakdown'))
                    logger.debug(i + ' : Files have been moved to Archive folder...')
                else:
                    # throw error as the child file is not exists.
                    try:
                        f = open(i.replace('PurchaseOrder', 'SizePerColourBreakdown'),encoding='utf-8',errors='ignore')
                    except FileNotFoundError:
                        logger.error('OrderNum: ' + str(orderNum) + ', SizePerColourBreakdown file not found!!!')
            except FileNotFoundError:
                logger.debug('File is not found!')


        else:
            orderNum = file_name.split('_', 2)[1]
            logger.debug('----------------------- Update Order Scenario -----------------------')
            logger.debug('Update Order Number: ' + str(orderNum))
            try:
                if get_file_validation(file_path, orderNum, 2):
                    # get the PurchaseOrder info and go down to the detail loop!!!!!!!!!!
                    order_object = get_purchase_order_first_page_text(i)
                    order_text = order_object.pages[0].extract_text_simple()
                    order_info_array = split_lines(order_text)
                    season = get_season(order_info_array[6])  # Column C
                    logger.debug('Season Code: ' + season)
                    folder_path =  os.path.join(excel_file_path,season)
                    old_sheet_pattern = str(orderNum) + '_Single_Sheet.*'
                    ###rename all the existed single excel sheet
                    rename_flag = rename_file_from_season(folder_path, old_sheet_pattern)

                    ###remove the order from summary excel sheet
                    ##logger.debug(os.path.join(excel_file_path,'Orders_Summary_' + season + '.xlsx'))
                    ##remove_flag = 0
                    ##df = pd.read_excel(os.path.join(excel_file_path,'Orders_Summary_' + season + '.xlsx')).iloc[:, 3]
                    ##min_row = 1000000000
                    ##max_row = 0
                    ##for j in range(0, len(df)):
                    ##    if str(df[j]) == str(orderNum):
                    ##        if j <= min_row:
                    ##            min_row = j
                    ##        if j >= max_row:
                    ##            max_row = j
                    ##logger.debug('min row: ' + str(min_row + 2))
                    ##logger.debug('max row: ' + str(max_row + 2))
                    ##if max_row != 2 and min_row != 1000000002:
                    ##    wb = openpyxl.load_workbook(os.path.join(excel_file_path,'Orders_Summary_' + season + '.xlsx'))
                    ##    wb_sheet_name = wb.get_sheet_by_name(wb.get_sheet_names()[0])
                        ###wb_sheet_name.delete_rows(min_row + 2, max_row + 2)
                    ##    wb_sheet_name.delete_rows(idx=min_row + 2, amount=max_row-min_row+1)  ## idx/amount ver2.0
                    ##    wb.save(os.path.join(excel_file_path,'Orders_Summary_' + season + '.xlsx'))
                    ##    wb.close()
                    ##    remove_flag = 1
                    ##else:
                    ##    logger.debug('Did not find this Order in the Orders Summary Excel Sheet...')
                    ##    remove_flag = 0

                    ###re-create the files
                    if rename_flag == 1:
                        brand = get_brand(order_info_array[1])  # Column A
                        order_type = get_order_type(order_info_array[0])  # Column B
                        order_no = get_order_no(order_info_array[3])  # Column D
                        department_no = get_department_no(order_info_array[3])  # Column E
                        product_no = get_product_no(order_info_array[3])  # Column F
                        date_of_order = get_date_of_order(order_info_array[5])  # Column G
                        product_name = get_product_name(order_info_array[4])  # Column H
                        development_no = get_development_no(order_info_array[9])  # Column I
                        no_of_pieces = get_no_of_pieces(order_info_array[12])  # Column J
                        download_date = datetime.date.today().strftime('%Y-%m-%d')  # Column V

                        ### create the price/country mapping information
                        terms_1st_position = 0
                        terms_last_position = 0
                        time_delivery_1st_position = 0
                        time_delivery_last_position = 0
                        price_1st_position = 0
                        price_last_position = 0
                        colourname_1st_position = 0
                        colourname_last_position = 0
                        for p in range(0, len(order_info_array)):
                            if re.search(r'Terms of Delivery', order_info_array[p]) is not None:
                                terms_1st_position = p
                            if re.search(r'Time of Delivery Planning Markets', order_info_array[p]) is not None:
                                terms_last_position = p - 1
                                time_delivery_1st_position = p
                            if re.search(r'Total: \d+', order_info_array[p]) is not None:
                                time_delivery_last_position = p - 1
                            if re.search(r'Invoice Average Price', order_info_array[p]) is not None:
                                price_1st_position = p
                            if (re.search(r'By accepting', order_info_array[p]) is not None) or (
                                    re.search(r'License Order', order_info_array[p]) is not None):
                                if price_last_position == 0:
                                    price_last_position = p - 1
                                elif price_last_position > p:
                                    price_last_position = p - 1
                            if (re.search(r'Article No H&M Colour Code', order_info_array[p])):
                                colourname_1st_position = p
                            if (re.search(r'Total Quantity:', order_info_array[p])):
                                colourname_last_position = p - 1

                        ### create the country/term mapping
                        term_dict = {}
                        term_dict.clear()
                        code_file_path = path_dicts['main_path']
                        term_dict = get_term_dicts(order_info_array, terms_1st_position + 1, terms_last_position + 1,country_codes)

                        ##create the time delivery mapping
                        time_delivery_dict = {}
                        time_delivery_dict.clear()
                        time_delivery_dict = get_delivery_dates_dicts(order_info_array, time_delivery_1st_position + 1,
                                                                      time_delivery_last_position + 1)
                        logger.debug('Print time Delivery Dict: ')
                        logger.debug(time_delivery_dict)

                        ##create price mapping
                        price_dict = {}
                        price_dict = get_price_dicts(order_info_array, price_1st_position + 1, price_last_position + 1)


                        ##create colourname mapping
                        colourname_dict = {}
                        colourname_dict.clear()
                        colourname_dict = get_colourname_dicts(order_info_array, colourname_1st_position + 1,
                                                               colourname_last_position + 1)


                        ###detail loop start here
                        detail_pages = []
                        detail_obj = get_sizecolourbreakdown_pages_text(
                            i.replace('PurchaseOrder', 'SizePerColourBreakdown'))
                        detail_pages = detail_obj.pages
                        for dp in range(0, len(detail_pages)):
                            country_name = ''  # Column K
                            fright_term = ''  # Column L
                            cost = ''  # Column R
                            currency = ''  # Column S
                            TOD = ''  # Column U
                            page_text = detail_pages[dp].extract_text_simple()
                            order_detail_array = split_lines(page_text)

                            assortment_1st_position = 0
                            assortment_last_position = 0
                            solid_1st_position = 0
                            solid_last_position = 0

                            for p in range(0, len(order_detail_array)):
                                if (re.search(r'Assortment', order_detail_array[p])):
                                    assortment_1st_position = p
                                if (re.search(r'Solid', order_detail_array[p])):
                                    solid_1st_position = p
                                    assortment_last_position = p - 1
                                if (re.search(r'Total', order_detail_array[p])):
                                    solid_last_position = p - 1

                            country_code = get_detail_country_code(order_detail_array[11])
                            ## get term

                            for k in term_dict.keys():
                                if re.search(country_code, str(k)) is not None:
                                    country_name = k
                                    fright_term = term_dict[k]
                                    break
                            ## get cost and currency
                            for k in price_dict.keys():
                                if re.search(country_code, str(k)) is not None:
                                    cost = price_dict[k].split('-')[0]
                                    currency = price_dict[k].split('-')[1]
                                    break

                            for k in time_delivery_dict.keys():
                                if re.search(country_code, str(k)) is not None:
                                    TOD = time_delivery_dict[k]
                                    break

                            ##get artical NO & colourcode+colourname
                            artical_list = re.findall(r'\d+.*\d+', order_detail_array[12])[0].split(' ')
                            colourcode_list = re.findall(r'\d+-.*', order_detail_array[13])[0].split(' ')
                            # print(country_name)
                            for a in artical_list:
                                artical_no = ''  # Column V
                                colourcode_colourname = ''  # Column M
                                size = ''  # Column N
                                l_no_assortment_qty = ''
                                l_assortment_qty = ''
                                packing_type = ''  # Column O
                                qty_artical = ''  # Column P
                                artical_no = a
                                for k in colourname_dict.keys():
                                    if a == k:
                                        colourcode_colourname = colourcode_list[artical_list.index(a)] + ' ' + str(
                                            colourname_dict[a])
                                        break
                                ## assorment/solid qty information

                                if assortment_1st_position > 0:
                                    packing_type = 'Assortment'
                                    # print(packing_type)
                                    no_of_asst_list = []
                                    for ap in range(assortment_1st_position + 1, assortment_last_position + 1):
                                        if re.search(r'No of Asst:', order_detail_array[ap]) is not None:
                                            no_of_asst_list = order_detail_array[ap].replace('No of Asst: ', '').split(
                                                ' ')
                                            break
                                    for ap in range(assortment_1st_position + 1, assortment_last_position + 1):
                                        if re.search(r'\(.*\)\*', order_detail_array[ap]) is not None:
                                            size = re.findall(r'\(.*\)\*', order_detail_array[ap])[0].replace(')*',
                                                                                                              '').replace(
                                                '(', '')
                                            if re.findall(r'\* \d+.*', order_detail_array[ap]) == []:
                                                next
                                            else:
                                                if len(re.findall(r'\* \d+.*', order_detail_array[ap])[0].replace('* ',
                                                                                                                  '').split(
                                                        ' ')) >= (artical_list.index(a)+1):
                                                    if re.findall(r'\* \d+.*', order_detail_array[ap]) is not [] and int(
                                                        re.findall(r'\* \d+.*', order_detail_array[ap])[0].replace('* ',
                                                                                                                   '').split(
                                                            ' ')[artical_list.index(a)]) > 0:
                                                        l_assortment_qty = int(
                                                            re.findall(r'\* \d+.*', order_detail_array[ap])[0].replace('* ',
                                                                                                                   '').split(
                                                            ' ')[artical_list.index(a)])
                                                        qty_artical = int(
                                                            re.findall(r'\* \d+.*', order_detail_array[ap])[0].replace('* ',
                                                                                                                   '').split(
                                                            ' ')[artical_list.index(a)]) * int(
                                                            no_of_asst_list[artical_list.index(a)])
                                                        l_no_assortment_qty = int(
                                                            no_of_asst_list[artical_list.index(a)])
                                                    ####call write_data_into_excel()
                                                    ##prepare dataset
                                                        data.append([brand, order_type, season, order_no, department_no,
                                                            product_no,
                                                            date_of_order, product_name, development_no, no_of_pieces,
                                                            country_name, fright_term.replace(',', ''), colourcode_colourname, size, packing_type, l_assortment_qty,l_no_assortment_qty,qty_artical, cost,
                                                            currency, int(no_of_pieces) * int(qty_artical), TOD,
                                                            artical_no,
                                                            download_date])
                                                    #write_data_into_excel(excel_file_path, season, orderNum, data)

                                if solid_1st_position > 0:
                                    packing_type = 'Solid'
                                    # print(packing_type)
                                    for sp in range(solid_1st_position + 1, solid_last_position + 1):
                                        if re.search(r'\(.*\)\*', order_detail_array[sp]) is not None:
                                            size = re.findall(r'\(.*\)\*', order_detail_array[sp])[0].replace(')*',
                                                                                                              '').replace(
                                                '(', '')
                                            # print(re.findall(r'\* \d+.*', order_detail_array[sp])[0].replace('* ',''))
                                            if re.findall(r'\* \d+.*', order_detail_array[sp]) == []:
                                                next
                                            else:
                                                if len(re.findall(r'\* \d+.*', order_detail_array[sp])[0].replace('* ',
                                                                                                                  '').split(
                                                        ' ')) >= (artical_list.index(a)+1):
                                                    if re.findall(r'\* \d+.*', order_detail_array[sp]) is not [] and int(
                                                        re.findall(r'\* \d+.*', order_detail_array[sp])[0].replace('* ',
                                                                                                                   '').split(
                                                            ' ')[artical_list.index(a)]) > 0:
                                                        l_assortment_qty = ''
                                                        l_no_assortment_qty = ''
                                                        qty_artical = int(
                                                        re.findall(r'\* \d+.*', order_detail_array[sp])[0].replace('* ',
                                                                                                                   '').split(
                                                            ' ')[artical_list.index(a)])
                                                    #######write data into excel sheet
                                                    ##read excel file , if file not existed, create a new one.
                                                    # prepare dataset
                                                        data.append([brand, order_type, season, order_no, department_no,
                                                            product_no,
                                                            date_of_order, product_name, development_no, no_of_pieces,
                                                            country_name, fright_term.replace(',', ''),
                                                            colourcode_colourname, size, packing_type, l_assortment_qty,l_no_assortment_qty,qty_artical,
                                                            cost,
                                                            currency, int(no_of_pieces) * int(qty_artical), TOD,
                                                            artical_no,
                                                            download_date])
                                                    ##write_data_into_excel(excel_file_path, season, orderNum, data)
                        detail_obj.close()
                        order_object.close()
                        ####end of size loop
                else:
                    # throw error as the child file is not exists.
                    # throw error as the child file is not exists.
                    try:
                        f = open(file_path + 'updated_' + str(orderNum) + '_SizePerColourBreakdown*',encoding='utf-8',errors='ignore')
                    except FileNotFoundError:
                        logger.error('OrderNum: ' + str(orderNum) + ', updated SizePerColourBreakdown file not found!!!')
                        exit(code='file is not found!!!')
            except FileNotFoundError:
                logger.error('OrderNum: ' + str(orderNum) + ', main exception!!!!')
                # move processed file into Archive folder
            if rename_flag == 1:
                logger.debug('Writing data into excel sheet...')
                write_data_into_excel(excel_file_path, season, orderNum, data)
                try:
                    if os.path.exists(i.replace('PurchaseOrder', 'SizePerColourBreakdown')):
                        ##get the PurchaseOrder info and go down to the detail loop!!!!!!!!!!
                        ####Update Order Scenarios
                        archive_path = path_dicts['archive_path']
                        shutil.move(i, archive_path + file_name)
                        shutil.move(i.replace('PurchaseOrder', 'SizePerColourBreakdown'),
                                    archive_path + file_name.replace('PurchaseOrder', 'SizePerColourBreakdown'))
                        logger.debug(i + ' : Files have been moved to Archive folder...')
                    else:
                        # throw error as the child file is not exists.,
                        try:
                            f = open(i.replace('PurchaseOrder', 'SizePerColourBreakdown'),encoding='utf-8',errors='ignore')
                        except FileNotFoundError:
                            logger.error(
                                'OrderNum: ' + str(orderNum) + ', updated SizePerColourBreakdown file not found!!!')
                except FileNotFoundError:
                    logger.error('File is not found!')
        # end of outer for loop

# See PyCharm help at https://www.jetbrains.com/help/pycharm/